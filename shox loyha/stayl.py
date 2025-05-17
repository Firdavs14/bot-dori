import asyncio
from aiogram import Bot, Dispatcher, F, types
from aiogram.enums import ParseMode
from aiogram.filters import CommandStart
from aiogram.types import (
    KeyboardButton, ReplyKeyboardMarkup, Message, ReplyKeyboardRemove
)
from aiogram.client.default import DefaultBotProperties
from openpyxl import load_workbook
from rapidfuzz import fuzz
import os
import re

TOKEN = "7781592878:AAG3k-S9KfBVCQf-sXD4gmKz9mR28gospwk"
GROUP_CHAT_ID = "-1002556986375"

dorilar = []
if os.path.exists("dorilar.xlsx"):
    wb = load_workbook("dorilar.xlsx")
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            dorilar.append({
                "nomi": str(row[0]).strip().lower(),
                "narxi": int(row[1])
            })
else:
    print("❌ dorilar.xlsx topilmadi!")

def lotin_to_kiril(text):
    replace_map = {
        "sh": "ш", "ch": "ч", "yo": "ё", "ya": "я", "yu": "ю", "ts": "ц",
        "a": "а", "b": "б", "d": "д", "e": "е", "f": "ф", "g": "г",
        "h": "ҳ", "i": "и", "j": "ж", "k": "к", "l": "л", "m": "м",
        "n": "н", "o": "о", "p": "п", "q": "қ", "r": "р", "s": "с",
        "t": "т", "u": "у", "v": "в", "x": "х", "y": "й", "z": "з",
        "ʼ": "ъ", "’": "ъ", "ʻ": "ъ"
    }
    for key in sorted(replace_map, key=len, reverse=True):
        text = re.sub(key, replace_map[key], text)
    return text

bot = Bot(token=TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher()

main_menu = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="🔍 Dori qidirish")],
        [KeyboardButton(text="📋 Retsept hisoblash")],
        [KeyboardButton(text="📍 Apteka joylashuvi")],
        [KeyboardButton(text="📞 Bog'lanish")],
    ],
    resize_keyboard=True
)

back_menu = ReplyKeyboardMarkup(
    keyboard=[[KeyboardButton(text="🔙 Ortga")]],
    resize_keyboard=True
)

@dp.message(CommandStart())
async def start(msg: Message):
    await msg.answer("👋 Xush kelibsiz! MED EKSPRES botiga.\nQuyidagilardan birini tanlang:", reply_markup=main_menu)

@dp.message(F.text == "📍 Apteka joylashuvi")
async def send_location(msg: Message):
    await msg.answer("📍 Manzil: <a href='https://maps.app.goo.gl/Q2ZtWd34w5rVxr9h6'>Google Maps'da ochish</a>")

@dp.message(F.text == "📞 Bog'lanish")
async def contact_info(msg: Message):
    await msg.answer("📞 Biz bilan bog‘lanish:   +998885570567")

@dp.message(F.text == "🔍 Dori qidirish")
async def dori_qidirish_start(msg: Message):
    await msg.answer("🔎 Dori nomining 3-4 harfini кирил ёки лотин алифбосида киритинг (масалан: пара, nur)", reply_markup=back_menu)

@dp.message(F.text == "📋 Retsept hisoblash")
async def retsept_start(msg: Message):
    await msg.answer("📸 Iltimos, retseptdagi dorilar ro‘yxatini yozing yoki rasm yuboring:", reply_markup=back_menu)

@dp.message(F.photo)
async def handle_photo(msg: Message):
    await msg.answer("🧮 Retsept hisoblanmoqda...")

@dp.message(F.text == "🔙 Ortga")
async def ortga(msg: Message):
    await msg.answer("🔙 Asosiy menyu", reply_markup=main_menu)

@dp.message()
async def process_text(msg: Message):
    text = msg.text.lower()
    if re.search(r'[a-z]', text):
        text = lotin_to_kiril(text)

    mos = []
    for item in dorilar:
        if fuzz.partial_ratio(text, item["nomi"]) > 80:
            mos.append(f"{item['nomi'].title()} - {item['narxi']} so'm")

    if mos:
        await msg.answer("🔎 Topilgan dorilar:\n" + "\n".join(mos), reply_markup=back_menu)
        await msg.answer(
            "<b>📍 Termiz Kardiologiya (Sapartak)</b>\n"
            "🏦 Xalq bank ko‘chasi\n"
            "📞 Apteka raqami: <a href='tel:+998885570567'>+998885570567</a>\n"
            "📲 Telegram: <a href='https://t.me/Big_Bull_Investi'>@Big_Bull_Investi</a>",
            disable_web_page_preview=True
        )
        await bot.send_message(GROUP_CHAT_ID, f"👤 @{msg.from_user.username or msg.from_user.full_name} dori qidirdi: {msg.text}")
    else:
        await msg.answer("❌ Afsuski, bu dori topilmadi. Iltimos, nomini to‘g‘ri kiriting.")

async def main():
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
